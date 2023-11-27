import os
import sys
from PIL import Image
from datetime import datetime
from openpyxl import Workbook
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl.styles import Alignment, Font
from pathlib import Path

# Get directory with files
sourceDirectory = sys.argv[1]  # directory with pictures
targetDirectory = sys.argv[2]  # directory where to save excel file
identifierForController = sys.argv[3]  # identirifer for ITE (optional but still need a string, like "")
directory = os.fsdecode(sourceDirectory)
filesCount = len(os.listdir(directory))
pictureFileCount = 0

# Get text and picture info from each file
eachTxtFile = []  # Intended to hold the txt file
eachFilesImage = []  #Intended to hold the image file
maxPictureWidth = 0
maxPictureHeight = 0

for file in os.listdir(directory):
    txtFileNAme = os.fsdecode(file)
    if txtFileNAme.endswith(".txt"):
        eachTxtFile.append(directory + "/" + txtFileNAme)
    elif txtFileNAme.endswith(".jpg") or txtFileNAme.endswith(".png") or txtFileNAme.endswith(".tiff"):
        pictureFileCount += 1
        # Pillow Image class
        with Image.open(directory + "/" + txtFileNAme) as image:
            eachFilesImage.append(directory + "/" + txtFileNAme)
            width, height = image.size
            if width > maxPictureWidth:
                maxPictureWidth = width
            elif height > maxPictureHeight:
                maxPictureHeight = height

# Initialise excel
wb = Workbook()
ws = wb.active

# Define some styles
ws.title = "Text-Image"
ws.column_dimensions['A'].width = 50
ws.column_dimensions['B'].width = maxPictureWidth * 0.1

myFont = Font(name='Courier',
              size=14,
              bold=False,
              italic=False,
              vertAlign=None,
              underline='none',
              strike=False,
              color='FF000000')

myAlignment = Alignment(horizontal='general',
                        vertical='top',
                        text_rotation=0,
                        wrap_text=True,
                        shrink_to_fit=False,
                        indent=0)

# Define and prepare the dictionary by matching filenames
picTextDictionary = {}

for txtFileName in eachTxtFile:
    pureImageFileName = txtFileName.split("/")
    pureFileName = Path(pureImageFileName[-1]).stem
    for imageFileName in eachFilesImage:
        pureImageFileName = imageFileName.split("/")
        pureImageFileName = Path(pureImageFileName[-1]).stem
        if (pureFileName == pureImageFileName):
            picTextDictionary[txtFileName] = imageFileName

eachText = list(picTextDictionary.keys())
eachImage = list(picTextDictionary.values())

# First column
for i in range(len(eachText)):
    # Set cell parameters/style
    ws.row_dimensions[i + 1].height = maxPictureHeight
    ws['A' + str(i + 1)].alignment = myAlignment
    ws['A' + str(i + 1)].font = myFont

    # Write text to cell
    text = open(eachText[i], 'r')
    ws['A' + str(i + 1)] = text.read()

# Second column
for i in range(len(eachImage)):
    img = OpenpyxlImage(eachImage[i])
    ws.add_image(img, 'B' + str(i + 1))

# Works
fileName = identifierForController + datetime.now().strftime('%Y-%m-%d %H:%M:%S') + ".xlsx"
wb.save(targetDirectory + "/" + fileName)
